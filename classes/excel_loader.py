from langchain.document_loaders.base import BaseLoader
from langchain.docstore.document import Document
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from PIL import Image as PILImage
from zipfile import BadZipFile
import io
import base64


class ExcelLoader(BaseLoader):
    def __init__(self, file_path):
        self.file_path = file_path

    def load(self):
        try:
            wb = load_workbook(self.file_path)
            documents = []

            for sheet in wb.sheetnames:
                ws = wb[sheet]
                content = f"Sheet: {sheet}\n\n"

                # テキストデータの抽出
                for row in ws.iter_rows(values_only=True):
                    content += " | ".join(str(cell) for cell in row if cell is not None) + "\n"

                # 画像の抽出
                for image in ws._images:
                    image.ref.seek(0)
                    img = PILImage.open(image.ref)
                    img_byte_arr = io.BytesIO()
                    img.save(img_byte_arr, format='PNG')
                    img_base64 = base64.b64encode(img_byte_arr.getvalue()).decode('utf-8')
                    content += f"\n[IMAGE:{img_base64}]\n"

                metadata = {
                    "source": self.file_path,
                    "sheet": sheet
                }
                doc = Document(page_content=content, metadata=metadata)
                documents.append(doc)

            return documents
        except InvalidFileException:
            print(f"{self.file_path} はパスワード保護されているか、無効なファイル形式です。スキップします。")
            raise
        except BadZipFile as e:
            print(f"Skipping file {self.file_path}: {e}")
            raise
        except Exception as e:
            print(f"Error loading file {self.file_path}: {e}")
            raise
        
if __name__ == '__main__':
    test = ExcelLoader(file_path='/data/5D_JHF_HPシステム基盤改更/100_アトミファイルサーバ/社外/COMSYS/030_要件定義/コンテンツ所管部署＆コンテンツベンダヒアリング/第1回ヒアリング/~$ヒアリングシート集計_171017.xlsx')
    test.load()